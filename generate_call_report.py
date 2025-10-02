#!/usr/bin/env python3
"""
Script para generar reportes en Excel de llamadas realizadas
Incluye datos completos de llamadas, transcripts, costos, y análisis
"""

import os
import sys
import asyncio
import argparse
from datetime import datetime, timezone, timedelta
from typing import Dict, Any, List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Add the app directory to the path
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app'))

from infrastructure.database_manager import DatabaseManager
from services.job_service import JobService
from services.account_service import AccountService
from config.settings import get_settings
from dotenv import load_dotenv

# Load environment variables
load_dotenv(os.path.join(os.path.dirname(__file__), 'app', '.env'))

class CallReportGenerator:
    """Generador de reportes de llamadas en Excel"""
    
    def __init__(self):
        self.settings = get_settings()
        self.db_manager = None
        self.job_service = None
        self.account_service = None
        
    async def initialize(self):
        """Inicializar conexiones a servicios"""
        self.db_manager = DatabaseManager(
            connection_string=os.getenv("MONGO_URI", "mongodb://localhost:27017"),
            database_name=os.getenv("MONGO_DB", "speechai_db")
        )
        await self.db_manager.connect()
        
        self.job_service = JobService(self.db_manager)
        self.account_service = AccountService(self.db_manager)
        
    async def get_calls_data(
        self, 
        account_id: Optional[str] = None,
        batch_id: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        status_filter: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        Obtener datos de llamadas con filtros
        
        Args:
            account_id: Filtrar por cuenta específica
            batch_id: Filtrar por batch específico
            start_date: Fecha de inicio
            end_date: Fecha de fin
            status_filter: Filtrar por estado específico
            
        Returns:
            Lista de diccionarios con datos de llamadas
        """
        print("📊 Extrayendo datos de llamadas...")
        
        # Construir query de filtros
        query = {}
        
        if account_id:
            query["account_id"] = account_id
            
        if batch_id:
            query["batch_id"] = batch_id
            
        if status_filter:
            query["status"] = status_filter
            
        # Filtro de fecha
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter["$gte"] = start_date
            if end_date:
                date_filter["$lte"] = end_date
            query["created_at"] = date_filter
        
        # Obtener jobs de la base de datos
        collection = self.db_manager.get_collection("jobs")
        cursor = collection.find(query).sort("created_at", -1)
        
        calls_data = []
        async for doc in cursor:
            # Extraer información básica
            call_data = self._extract_call_data(doc)
            calls_data.append(call_data)
            
        print(f"✅ {len(calls_data)} llamadas encontradas")
        return calls_data
    
    def _extract_call_data(self, doc: Dict[str, Any]) -> Dict[str, Any]:
        """Extraer y estructurar datos de una llamada"""
        
        # Información básica del job
        job_id = str(doc.get("_id", ""))
        account_id = doc.get("account_id", "")
        batch_id = doc.get("batch_id", "")
        status = doc.get("status", "unknown")
        
        # Información del contacto
        contact_name = doc.get("nombre", "")
        contact_rut = doc.get("rut", "")
        phone_number = doc.get("to_number", "")
        debt_amount = doc.get("deuda", 0)
        
        # Fechas
        created_at = doc.get("created_at")
        started_at = doc.get("call_started_at")
        ended_at = doc.get("call_ended_at")
        finished_at = doc.get("finished_at")
        
        # Información de la llamada
        call_id = doc.get("call_id", "")
        call_duration_seconds = doc.get("call_duration_seconds", 0)
        call_duration_minutes = round(call_duration_seconds / 60, 2) if call_duration_seconds else 0
        
        # Resultado de la llamada
        call_result = doc.get("call_result", {})
        call_success = call_result.get("success", False)
        call_status = "No iniciada"
        call_cost = 0
        transcript = ""
        recording_url = ""
        public_log_url = ""
        disconnection_reason = ""
        
        # Variables capturadas
        collected_vars = {}
        fecha_pago_cliente = ""
        monto_pago_cliente = ""
        
        if call_result:
            call_status = call_result.get("status", "unknown")
            
            # Información de costos
            summary = call_result.get("summary", {})
            if summary.get("call_cost"):
                cost_data = summary["call_cost"]
                if isinstance(cost_data, dict):
                    call_cost = cost_data.get("combined_cost", 0)
                else:
                    call_cost = float(cost_data) if cost_data else 0
            
            # URLs y datos adicionales
            transcript = summary.get("transcript", "")
            recording_url = summary.get("recording_url", "")
            public_log_url = summary.get("public_log_url", "")
            disconnection_reason = summary.get("disconnection_reason", "")
            
            # Variables dinámicas capturadas
            collected_vars = summary.get("collected_dynamic_variables", {})
            fecha_pago_cliente = collected_vars.get("fecha_pago_cliente", "")
            monto_pago_cliente = collected_vars.get("monto_pago_cliente", "")
        
        # Calcular duración legible
        duration_display = self._format_duration(call_duration_seconds)
        
        # Determinar resultado final
        final_result = self._determine_final_result(status, call_success, call_status)
        
        return {
            # Identificadores
            "job_id": job_id,
            "account_id": account_id,
            "batch_id": batch_id,
            "call_id": call_id,
            
            # Información del contacto
            "contact_name": contact_name,
            "contact_rut": contact_rut,
            "phone_number": phone_number,
            "debt_amount": debt_amount,
            
            # Estado y resultado
            "job_status": status,
            "call_status": call_status,
            "final_result": final_result,
            "success": call_success,
            "disconnection_reason": disconnection_reason,
            
            # Fechas y tiempos
            "created_at": created_at,
            "started_at": started_at,
            "ended_at": ended_at,
            "finished_at": finished_at,
            "duration_seconds": call_duration_seconds,
            "duration_minutes": call_duration_minutes,
            "duration_display": duration_display,
            
            # Costos
            "call_cost": call_cost,
            
            # Contenido de la llamada
            "transcript": transcript,
            "recording_url": recording_url,
            "public_log_url": public_log_url,
            
            # Variables capturadas
            "fecha_pago_cliente": fecha_pago_cliente,
            "monto_pago_cliente": monto_pago_cliente,
            "collected_variables": str(collected_vars) if collected_vars else "",
        }
    
    def _format_duration(self, seconds: int) -> str:
        """Formatear duración en formato legible"""
        if not seconds:
            return "0:00"
            
        minutes = seconds // 60
        secs = seconds % 60
        return f"{minutes}:{secs:02d}"
    
    def _determine_final_result(self, job_status: str, call_success: bool, call_status: str) -> str:
        """Determinar el resultado final de la llamada"""
        if job_status == "failed":
            return "Falló"
        elif job_status == "suspended":
            return "Suspendido"
        elif job_status == "pending":
            return "Pendiente"
        elif job_status in ["completed", "done"]:
            if call_success:
                return "Exitosa"
            else:
                return f"Completada - {call_status}"
        else:
            return f"En progreso - {call_status}"
    
    def generate_excel_report(
        self, 
        calls_data: List[Dict[str, Any]], 
        filename: str,
        include_summary: bool = True
    ) -> str:
        """
        Generar reporte Excel con los datos de llamadas
        
        Args:
            calls_data: Lista de datos de llamadas
            filename: Nombre del archivo a generar
            include_summary: Si incluir hoja de resumen
            
        Returns:
            Ruta del archivo generado
        """
        print(f"📝 Generando reporte Excel: {filename}")
        
        # Crear DataFrame principal
        df_calls = pd.DataFrame(calls_data)
        
        # Ordenar columnas para mejor legibilidad
        column_order = [
            "job_id", "account_id", "batch_id", "contact_name", "contact_rut", 
            "phone_number", "debt_amount", "final_result", "job_status", 
            "call_status", "success", "duration_display", "duration_minutes",
            "call_cost", "created_at", "started_at", "ended_at", "finished_at",
            "disconnection_reason", "fecha_pago_cliente", "monto_pago_cliente",
            "transcript", "recording_url", "public_log_url", "call_id", "collected_variables"
        ]
        
        # Reordenar columnas (solo las que existen)
        existing_columns = [col for col in column_order if col in df_calls.columns]
        df_calls = df_calls[existing_columns]
        
        # Renombrar columnas para mejor legibilidad
        column_names = {
            "job_id": "ID Job",
            "account_id": "Cuenta",
            "batch_id": "Lote",
            "contact_name": "Nombre Contacto",
            "contact_rut": "RUT/DNI",
            "phone_number": "Teléfono",
            "debt_amount": "Monto Deuda",
            "final_result": "Resultado Final",
            "job_status": "Estado Job",
            "call_status": "Estado Llamada",
            "success": "Éxito",
            "duration_display": "Duración",
            "duration_minutes": "Minutos",
            "call_cost": "Costo (USD)",
            "created_at": "Creado",
            "started_at": "Iniciado",
            "ended_at": "Terminado",
            "finished_at": "Finalizado",
            "disconnection_reason": "Motivo Desconexión",
            "fecha_pago_cliente": "Fecha Pago Prometida",
            "monto_pago_cliente": "Monto Pago Prometido",
            "transcript": "Transcripción",
            "recording_url": "URL Grabación",
            "public_log_url": "URL Log Público",
            "call_id": "ID Llamada Retell",
            "collected_variables": "Variables Capturadas"
        }
        
        df_calls.rename(columns=column_names, inplace=True)
        
        # Crear archivo Excel con openpyxl para mejor control
        wb = Workbook()
        
        # Hoja principal de llamadas
        ws_calls = wb.active
        ws_calls.title = "Llamadas Detalladas"
        
        # Agregar datos al worksheet
        for r in dataframe_to_rows(df_calls, index=False, header=True):
            ws_calls.append(r)
        
        # Aplicar formato a la hoja de llamadas
        self._format_calls_sheet(ws_calls, len(df_calls))
        
        # Generar hoja de resumen si se solicita
        if include_summary:
            ws_summary = wb.create_sheet("Resumen")
            self._generate_summary_sheet(ws_summary, calls_data)
        
        # Guardar archivo
        wb.save(filename)
        print(f"✅ Reporte generado: {filename}")
        
        return filename
    
    def _format_calls_sheet(self, ws, num_rows: int):
        """Aplicar formato a la hoja de llamadas"""
        
        # Configurar anchos de columnas
        column_widths = {
            'A': 15,  # ID Job
            'B': 12,  # Cuenta
            'C': 20,  # Lote
            'D': 25,  # Nombre
            'E': 15,  # RUT
            'F': 15,  # Teléfono
            'G': 12,  # Deuda
            'H': 15,  # Resultado
            'I': 12,  # Estado Job
            'J': 15,  # Estado Llamada
            'K': 8,   # Éxito
            'L': 10,  # Duración
            'M': 10,  # Minutos
            'N': 10,  # Costo
            'O': 18,  # Creado
            'P': 18,  # Iniciado
            'Q': 18,  # Terminado
            'R': 18,  # Finalizado
            'S': 20,  # Motivo
            'T': 15,  # Fecha Pago
            'U': 15,  # Monto Pago
            'V': 50,  # Transcript
            'W': 30,  # URL Grabación
            'X': 30,  # URL Log
            'Y': 25,  # Call ID
            'Z': 30,  # Variables
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Formato del header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Aplicar bordes a toda la tabla
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=num_rows + 1):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # No aplicar a headers
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Crear tabla
        if num_rows > 0:
            from openpyxl.utils import get_column_letter
            max_col = ws.max_column
            max_col_letter = get_column_letter(max_col)
            
            table = Table(
                displayName="CallsTable",
                ref=f"A1:{max_col_letter}{num_rows + 1}"
            )
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
    
    def _generate_summary_sheet(self, ws, calls_data: List[Dict[str, Any]]):
        """Generar hoja de resumen con estadísticas"""
        
        # Título
        ws['A1'] = "RESUMEN DE LLAMADAS"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:D1')
        
        # Estadísticas generales
        total_calls = len(calls_data)
        successful_calls = len([c for c in calls_data if c.get('success')])
        failed_calls = total_calls - successful_calls
        success_rate = (successful_calls / total_calls * 100) if total_calls > 0 else 0
        
        total_cost = sum(c.get('call_cost', 0) for c in calls_data)
        total_minutes = sum(c.get('duration_minutes', 0) for c in calls_data)
        avg_duration = total_minutes / max(successful_calls, 1)
        
        # Agregar estadísticas
        row = 3
        stats = [
            ("Total de Llamadas:", total_calls),
            ("Llamadas Exitosas:", successful_calls),
            ("Llamadas Fallidas:", failed_calls),
            ("Tasa de Éxito:", f"{success_rate:.1f}%"),
            ("Costo Total:", f"${total_cost:.2f} USD"),
            ("Tiempo Total:", f"{total_minutes:.1f} minutos"),
            ("Duración Promedio:", f"{avg_duration:.1f} minutos"),
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Resumen por estado
        row += 2
        ws[f'A{row}'] = "RESUMEN POR ESTADO"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        status_summary = {}
        for call in calls_data:
            status = call.get('final_result', 'Desconocido')
            status_summary[status] = status_summary.get(status, 0) + 1
        
        ws[f'A{row}'] = "Estado"
        ws[f'B{row}'] = "Cantidad"
        ws[f'C{row}'] = "Porcentaje"
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
            cell.font = Font(bold=True)
        row += 1
        
        for status, count in status_summary.items():
            percentage = (count / total_calls * 100) if total_calls > 0 else 0
            ws[f'A{row}'] = status
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{percentage:.1f}%"
            row += 1
        
        # Ajustar anchos de columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

async def main():
    """Función principal del script"""
    
    parser = argparse.ArgumentParser(description="Generar reporte Excel de llamadas")
    parser.add_argument("--account-id", help="Filtrar por cuenta específica")
    parser.add_argument("--batch-id", help="Filtrar por batch específico")
    parser.add_argument("--start-date", help="Fecha de inicio (YYYY-MM-DD)")
    parser.add_argument("--end-date", help="Fecha de fin (YYYY-MM-DD)")
    parser.add_argument("--status", help="Filtrar por estado (pending, completed, failed, etc.)")
    parser.add_argument("--output", default="call_report.xlsx", help="Nombre del archivo de salida")
    parser.add_argument("--no-summary", action="store_true", help="No incluir hoja de resumen")
    parser.add_argument("--days", type=int, help="Últimos N días (alternativa a start-date)")
    
    args = parser.parse_args()
    
    # Configurar fechas
    start_date = None
    end_date = None
    
    if args.days:
        end_date = datetime.now(timezone.utc)
        start_date = end_date - timedelta(days=args.days)
    elif args.start_date:
        start_date = datetime.fromisoformat(args.start_date).replace(tzinfo=timezone.utc)
    if args.end_date:
        end_date = datetime.fromisoformat(args.end_date).replace(tzinfo=timezone.utc)
    
    # Generar nombre de archivo con timestamp si no se especifica
    if args.output == "call_report.xlsx":
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = f"call_report_{timestamp}.xlsx"
    
    print("🚀 Iniciando generación de reporte de llamadas")
    print("=" * 50)
    
    if args.account_id:
        print(f"📊 Cuenta: {args.account_id}")
    if args.batch_id:
        print(f"📦 Batch: {args.batch_id}")
    if start_date:
        print(f"📅 Desde: {start_date.strftime('%Y-%m-%d')}")
    if end_date:
        print(f"📅 Hasta: {end_date.strftime('%Y-%m-%d')}")
    if args.status:
        print(f"🏷️ Estado: {args.status}")
    
    print(f"💾 Archivo: {args.output}")
    print("=" * 50)
    
    try:
        # Inicializar generador
        generator = CallReportGenerator()
        await generator.initialize()
        
        # Obtener datos
        calls_data = await generator.get_calls_data(
            account_id=args.account_id,
            batch_id=args.batch_id,
            start_date=start_date,
            end_date=end_date,
            status_filter=args.status
        )
        
        if not calls_data:
            print("⚠️ No se encontraron llamadas con los filtros especificados")
            return
        
        # Generar reporte
        report_file = generator.generate_excel_report(
            calls_data, 
            args.output,
            include_summary=not args.no_summary
        )
        
        print("=" * 50)
        print(f"✅ Reporte generado exitosamente: {report_file}")
        print(f"📊 Total de llamadas incluidas: {len(calls_data)}")
        
        # Mostrar estadísticas rápidas
        successful = len([c for c in calls_data if c.get('success')])
        success_rate = (successful / len(calls_data) * 100) if calls_data else 0
        total_cost = sum(c.get('call_cost', 0) for c in calls_data)
        
        print(f"🎯 Tasa de éxito: {success_rate:.1f}%")
        print(f"💰 Costo total: ${total_cost:.2f} USD")
        
    except Exception as e:
        print(f"❌ Error generando reporte: {e}")
        raise
    
    finally:
        # Cleanup
        if generator.db_manager:
            await generator.db_manager.close()

if __name__ == "__main__":
    # Verificar dependencias
    try:
        import pandas as pd
        import openpyxl
    except ImportError as e:
        print("❌ Dependencias faltantes. Instalar con:")
        print("pip install pandas openpyxl")
        sys.exit(1)
    
    asyncio.run(main())